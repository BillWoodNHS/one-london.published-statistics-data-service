"""Tests for normalize_payload_to_csv, _all_csvs_from_zip, _excel_to_csv,
_count_csv_rows, and build_artifact in function_app/src/download_and_normalize.py.
"""

from __future__ import annotations

import io
import zipfile

import openpyxl
import pandas as pd
import pytest

from function_app.src.download_and_normalize import (
    _all_csvs_from_zip,
    _count_csv_rows,
    _excel_to_csv,
    _ods_to_csv,
    _unpivot,
    normalize_payload_to_csv,
    resolve_sub_table_adls_prefix,
)
from function_app.src.models import SubTableConfig, TargetConfig, UnpivotConfig

# ---------------------------------------------------------------------------
# Helpers to build test payloads in memory
# ---------------------------------------------------------------------------


def _make_zip(*members: tuple[str, bytes]) -> bytes:
    """Build an in-memory ZIP containing (name, data) members."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for name, data in members:
            zf.writestr(name, data)
    return buf.getvalue()


def _make_xlsx(headers: list[str], rows: list[list]) -> bytes:
    """Build an in-memory .xlsx workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_ods(headers: list[str], rows: list[list]) -> bytes:
    """Build an in-memory .ods workbook."""
    frame = pd.DataFrame(rows, columns=headers)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="odf") as writer:
        frame.to_excel(writer, index=False)
    return buf.getvalue()


CSV_SIMPLE = b"col_a,col_b\n1,2\n3,4\n5,6\n"
CSV_HEADER_ONLY = b"col_a,col_b\n"
CSV_EMPTY = b""


# ---------------------------------------------------------------------------
# _count_csv_rows
# ---------------------------------------------------------------------------


class TestCountCsvRows:
    def test_counts_data_rows_excluding_header(self):
        assert _count_csv_rows(CSV_SIMPLE) == 3

    def test_header_only_returns_zero(self):
        assert _count_csv_rows(CSV_HEADER_ONLY) == 0

    def test_empty_payload_returns_zero(self):
        assert _count_csv_rows(CSV_EMPTY) == 0

    def test_bom_encoded_utf8(self):
        bom_csv = b"\xef\xbb\xbfcol_a,col_b\n1,2\n"
        assert _count_csv_rows(bom_csv) == 1

    def test_latin1_encoded(self):
        latin1_csv = "col\n\xe9l\xe8ve\n".encode("latin-1")
        assert _count_csv_rows(latin1_csv) == 1


# ---------------------------------------------------------------------------
# _all_csvs_from_zip
# ---------------------------------------------------------------------------


class TestAllCsvsFromZip:
    def test_extracts_csv_from_zip(self):
        payload = _make_zip(("data.csv", CSV_SIMPLE))
        ((name, content, metrics),) = _all_csvs_from_zip(payload)
        assert name == "data.csv"
        assert content == CSV_SIMPLE
        assert metrics["source_file_type"] == "zip"
        assert metrics["extracted_from_archive"] is True
        assert metrics["converted_to_csv"] is False
        assert metrics["raw_row_count"] == 3

    def test_extracts_all_files_from_zip(self):
        xlsx_payload = _make_xlsx(["a", "b"], [[1, 2]])
        payload = _make_zip(
            ("spreadsheet.xlsx", xlsx_payload),
            ("data.csv", CSV_SIMPLE),
        )
        results = _all_csvs_from_zip(payload)
        names = {name for name, _, _ in results}
        assert names == {"spreadsheet.csv", "data.csv"}

        by_name = {name: metrics for name, _, metrics in results}
        assert by_name["spreadsheet.csv"]["converted_to_csv"] is True
        assert by_name["data.csv"]["converted_to_csv"] is False

    def test_converts_xlsx_when_no_csv_in_zip(self):
        xlsx_payload = _make_xlsx(["x", "y"], [[10, 20], [30, 40]])
        payload = _make_zip(("report.xlsx", xlsx_payload))
        ((name, content, metrics),) = _all_csvs_from_zip(payload)
        assert name == "report.csv"
        assert metrics["converted_to_csv"] is True
        assert metrics["source_file_type"] == "zip"
        assert metrics["extracted_from_archive"] is True
        assert metrics["raw_row_count"] == 2

    def test_raises_when_zip_contains_no_usable_file(self):
        payload = _make_zip(("readme.txt", b"some text"))
        with pytest.raises(ValueError, match="ZIP did not contain"):
            _all_csvs_from_zip(payload)

    def test_converts_ods_when_no_csv_or_excel_in_zip(self):
        ods_payload = _make_ods(["x", "y"], [[10, 20], [30, 40]])
        payload = _make_zip(("report.ods", ods_payload))
        ((name, content, metrics),) = _all_csvs_from_zip(payload)
        assert name == "report.csv"
        assert metrics["converted_to_csv"] is True
        assert metrics["source_file_type"] == "zip"
        assert metrics["extracted_from_archive"] is True
        assert metrics["raw_row_count"] == 2


# ---------------------------------------------------------------------------
# _unpivot
# ---------------------------------------------------------------------------


class TestUnpivot:
    def test_melts_non_id_columns_into_long_format(self):
        df = pd.DataFrame(
            {
                "Org Code": ["A1", "B1"],
                "Apr 2025": [10, 20],
                "May 2025": [11, 21],
            }
        )
        config = UnpivotConfig(
            id_columns=["Org Code"],
            variable_column_name="reporting_period",
            value_column_name="value",
        )
        result = _unpivot(df, config)
        assert list(result.columns) == ["Org Code", "reporting_period", "value"]
        assert len(result.index) == 4
        rows = {
            (r["Org Code"], r["reporting_period"], r["value"])
            for _, r in result.iterrows()
        }
        assert rows == {
            ("A1", "Apr 2025", 10),
            ("A1", "May 2025", 11),
            ("B1", "Apr 2025", 20),
            ("B1", "May 2025", 21),
        }

    def test_is_generic_for_metric_per_column_not_just_dates(self):
        # The reshape is purely structural — it should behave identically
        # whether the melted headers are reporting periods or metric names.
        df = pd.DataFrame(
            {
                "Org Code": ["A1"],
                "Attendances": [100],
                "Admissions": [40],
            }
        )
        config = UnpivotConfig(
            id_columns=["Org Code"],
            variable_column_name="metric",
            value_column_name="value",
        )
        result = _unpivot(df, config)
        assert list(result.columns) == ["Org Code", "metric", "value"]
        assert set(result["metric"]) == {"Attendances", "Admissions"}

    def test_missing_id_column_raises_clear_error(self):
        df = pd.DataFrame({"Org Code": ["A1"], "Apr 2025": [10]})
        config = UnpivotConfig(
            id_columns=["Org Code", "Org name"],
            variable_column_name="reporting_period",
        )
        with pytest.raises(ValueError, match="Org name"):
            _unpivot(df, config)


# ---------------------------------------------------------------------------
# _excel_to_csv
# ---------------------------------------------------------------------------


class TestExcelToCsv:
    def test_converts_xlsx_to_csv(self):
        xlsx = _make_xlsx(["name", "score"], [["alice", 95], ["bob", 87]])
        ((name, content, metrics),) = _excel_to_csv("my_report", xlsx, ".xlsx")
        assert name == "my_report.csv"
        decoded = content.decode("utf-8")
        assert "name,score" in decoded
        assert "alice" in decoded
        assert metrics["converted_to_csv"] is True
        assert metrics["source_file_type"] == "excel"
        assert metrics["raw_row_count"] == 2
        assert metrics["normalized_row_count"] == 2
        assert metrics["extracted_from_archive"] is False

    def test_converted_csv_has_correct_row_count(self):
        xlsx = _make_xlsx(["a"], [[1], [2], [3], [4], [5]])
        ((_, content, metrics),) = _excel_to_csv("data", xlsx, ".xlsx")
        assert metrics["raw_row_count"] == 5
        assert _count_csv_rows(content) == 5

    def test_excel_sheet_selector_reads_named_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        wb.active.append(["junk"])
        data_sheet = wb.create_sheet("Data")
        data_sheet.append(["a", "b"])
        data_sheet.append([1, 2])
        buf = io.BytesIO()
        wb.save(buf)
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            excel_sheet="Data",
        )
        ((name, content, metrics),) = _excel_to_csv(
            "report", buf.getvalue(), ".xlsx", target
        )
        assert name == "report.csv"
        decoded = content.decode("utf-8")
        assert "a,b" in decoded
        assert metrics["raw_row_count"] == 1

    def test_sheet_splitting_produces_one_output_per_matched_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        wb.active.append(["ignore me"])
        monthly_jan = wb.create_sheet("Monthly Jan")
        monthly_jan.append(["a", "b"])
        monthly_jan.append([1, 2])
        monthly_feb = wb.create_sheet("Monthly Feb")
        monthly_feb.append(["a", "b"])
        monthly_feb.append([3, 4])
        buf = io.BytesIO()
        wb.save(buf)

        sub_table = SubTableConfig(
            object_name_suffix="MONTHLY",
            adls_path_prefix="dataset/monthly",
            sheet_name_patterns=["^Monthly"],
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            sub_tables=[sub_table],
        )
        results = _excel_to_csv("report", buf.getvalue(), ".xlsx", target)
        assert len(results) == 2
        names = {name for name, _, _ in results}
        assert names == {"report__Monthly_Jan.csv", "report__Monthly_Feb.csv"}
        for _, _, metrics in results:
            assert metrics["matched_sub_table_object_name_suffix"] == "MONTHLY"
            assert metrics["matched_sub_table_adls_path_prefix"] == "dataset/monthly"
            assert metrics["source_sheet_name"] in {"Monthly Jan", "Monthly Feb"}

    def test_sheet_splitting_matches_any_of_multiple_patterns(self):
        # Real-world sheets vary in wording across releases (e.g. "national
        # ... data" vs "national ... cases") — sheet_name_patterns is a list
        # so any alternative can match.
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        wb.active.append(["ignore me"])
        national = wb.create_sheet("National Cases")
        national.append(["a", "b"])
        national.append([1, 2])
        buf = io.BytesIO()
        wb.save(buf)

        sub_table = SubTableConfig(
            object_name_suffix="NATIONAL",
            adls_path_prefix="dataset/national",
            sheet_name_patterns=["national.*data", "national.*cases"],
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            sub_tables=[sub_table],
        )
        results = _excel_to_csv("report", buf.getvalue(), ".xlsx", target)
        assert len(results) == 1
        assert results[0][2]["source_sheet_name"] == "National Cases"

    def test_sheet_splitting_with_start_cell_skips_header_offset(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["B3"] = "a"
        ws["C3"] = "b"
        ws["B4"] = 1
        ws["C4"] = 2
        buf = io.BytesIO()
        wb.save(buf)

        sub_table = SubTableConfig(
            object_name_suffix="DATA",
            adls_path_prefix="dataset/data",
            sheet_name_patterns=["Data"],
            start_cell="B3",
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            sub_tables=[sub_table],
        )
        ((name, content, metrics),) = _excel_to_csv(
            "report", buf.getvalue(), ".xlsx", target
        )
        decoded = content.decode("utf-8")
        assert "a,b" in decoded
        assert "1,2" in decoded
        assert metrics["raw_row_count"] == 1

    def test_sheet_splitting_with_unpivot_reshapes_wide_period_columns(self):
        # Mirrors a CDI-style "Table_1_national_data" sheet: id columns
        # followed by one column per reporting month.
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        wb.active.append(["junk"])
        national = wb.create_sheet("Table_1_national_data")
        national.append(["Org Code", "Org name", "Apr 2025", "May 2025"])
        national.append(["A1", "Alpha", 10, 11])
        national.append(["B1", "Beta", 20, 21])
        buf = io.BytesIO()
        wb.save(buf)

        sub_table = SubTableConfig(
            object_name_suffix="NATIONAL",
            adls_path_prefix="dataset/national",
            sheet_name_patterns=["Table_1_national_data"],
            unpivot=UnpivotConfig(
                id_columns=["Org Code", "Org name"],
                variable_column_name="reporting_period",
                value_column_name="value",
            ),
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            sub_tables=[sub_table],
        )
        ((name, content, metrics),) = _excel_to_csv(
            "report", buf.getvalue(), ".xlsx", target
        )
        decoded = content.decode("utf-8")
        assert "Org Code,Org name,reporting_period,value" in decoded
        assert metrics["raw_row_count"] == 4
        assert "Apr 2025" in decoded and "May 2025" in decoded

    def test_sheet_splitting_raises_when_no_sheet_matches(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        wb.active.append(["junk"])
        buf = io.BytesIO()
        wb.save(buf)

        sub_table = SubTableConfig(
            object_name_suffix="DATA",
            adls_path_prefix="dataset/data",
            sheet_name_patterns=["^Data"],
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            sub_tables=[sub_table],
        )
        with pytest.raises(ValueError, match="None of the configured"):
            _excel_to_csv("report", buf.getvalue(), ".xlsx", target)


class TestOdsToCsv:
    def test_converts_ods_to_csv(self):
        ods = _make_ods(["name", "score"], [["alice", 95], ["bob", 87]])
        ((name, content, metrics),) = _ods_to_csv("my_report", ods, ".ods")
        assert name == "my_report.csv"
        decoded = content.decode("utf-8")
        assert "name,score" in decoded
        assert "alice" in decoded
        assert metrics["converted_to_csv"] is True
        assert metrics["source_file_type"] == "ods"
        assert metrics["raw_row_count"] == 2
        assert metrics["normalized_row_count"] == 2
        assert metrics["extracted_from_archive"] is False

    def test_converted_ods_has_correct_row_count(self):
        ods = _make_ods(["a"], [[1], [2], [3], [4], [5]])
        ((_, content, metrics),) = _ods_to_csv("data", ods, ".ods")
        assert metrics["raw_row_count"] == 5
        assert _count_csv_rows(content) == 5

    def test_empty_ods_returns_zero_rows(self):
        ods = _make_ods(["header"], [])
        ((_, content, metrics),) = _ods_to_csv("empty", ods, ".ods")
        assert metrics["raw_row_count"] == 0
        assert _count_csv_rows(content) == 0


# ---------------------------------------------------------------------------
# normalize_payload_to_csv
# ---------------------------------------------------------------------------


class TestNormalizePayloadToCsv:
    def test_csv_passthrough(self):
        ((name, content, h, metrics),) = normalize_payload_to_csv(
            "data.csv", CSV_SIMPLE
        )
        assert name == "data.csv"
        assert content == CSV_SIMPLE
        assert metrics["source_file_type"] == "csv"
        assert metrics["converted_to_csv"] is False
        assert metrics["extracted_from_archive"] is False
        assert len(h) == 64  # SHA-256 hex

    def test_zip_containing_csv(self):
        zip_payload = _make_zip(("result.csv", CSV_SIMPLE))
        ((name, content, h, metrics),) = normalize_payload_to_csv(
            "archive.zip", zip_payload
        )
        assert name == "result.csv"
        assert metrics["source_file_type"] == "zip"
        assert metrics["extracted_from_archive"] is True
        assert metrics["converted_to_csv"] is False

    def test_zip_containing_multiple_files(self):
        xlsx = _make_xlsx(["a", "b"], [[1, 2]])
        zip_payload = _make_zip(
            ("report.xlsx", xlsx),
            ("data.csv", CSV_SIMPLE),
        )
        results = normalize_payload_to_csv("archive.zip", zip_payload)
        names = {name for name, _, _, _ in results}
        assert names == {"report.csv", "data.csv"}

    def test_zip_containing_xlsx(self):
        xlsx = _make_xlsx(["a", "b"], [[1, 2]])
        zip_payload = _make_zip(("report.xlsx", xlsx))
        ((name, content, h, metrics),) = normalize_payload_to_csv(
            "archive.zip", zip_payload
        )
        assert name == "report.csv"
        assert metrics["converted_to_csv"] is True
        assert metrics["source_file_type"] == "zip"

    def test_bare_xlsx(self):
        xlsx = _make_xlsx(["col"], [["val1"], ["val2"]])
        ((name, content, h, metrics),) = normalize_payload_to_csv("report.xlsx", xlsx)
        assert name == "report.csv"
        assert metrics["source_file_type"] == "excel"
        assert metrics["converted_to_csv"] is True
        assert metrics["raw_row_count"] == 2

    def test_bare_xls_extension_handled(self):
        # Build a valid xlsx and lie about the extension
        xlsx = _make_xlsx(["x"], [[1]])
        ((name, content, h, metrics),) = normalize_payload_to_csv("old.xls", xlsx)
        assert name == "old.csv"

    def test_bare_ods(self):
        ods = _make_ods(["col"], [["val1"], ["val2"]])
        ((name, content, h, metrics),) = normalize_payload_to_csv("report.ods", ods)
        assert name == "report.csv"
        assert metrics["source_file_type"] == "ods"
        assert metrics["converted_to_csv"] is True
        assert metrics["raw_row_count"] == 2

    def test_zip_containing_ods(self):
        ods = _make_ods(["a", "b"], [[1, 2]])
        zip_payload = _make_zip(("report.ods", ods))
        ((name, content, h, metrics),) = normalize_payload_to_csv(
            "archive.zip", zip_payload
        )
        assert name == "report.csv"
        assert metrics["converted_to_csv"] is True
        assert metrics["source_file_type"] == "zip"

    def test_unsupported_extension_raises(self):
        with pytest.raises(ValueError, match="Unsupported file type"):
            normalize_payload_to_csv("report.pdf", b"%PDF-1.4")

    def test_csv_with_unpivot_config_reshapes_wide_columns(self):
        csv_payload = (
            b"Org Code,Org name,Apr 2025,May 2025\nA1,Alpha,10,11\nB1,Beta,20,21\n"
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            unpivot=UnpivotConfig(
                id_columns=["Org Code", "Org name"],
                variable_column_name="reporting_period",
                value_column_name="value",
            ),
        )
        ((name, content, h, metrics),) = normalize_payload_to_csv(
            "data.csv", csv_payload, target
        )
        assert name == "data.csv"
        decoded = content.decode("utf-8")
        assert "Org Code,Org name,reporting_period,value" in decoded
        assert metrics["converted_to_csv"] is True
        assert metrics["raw_row_count"] == 2
        assert metrics["normalized_row_count"] == 4
        assert len(h) == 64

    def test_csv_without_unpivot_config_still_passes_through(self):
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
        )
        ((name, content, h, metrics),) = normalize_payload_to_csv(
            "data.csv", CSV_SIMPLE, target
        )
        assert content == CSV_SIMPLE
        assert metrics["converted_to_csv"] is False

    def test_reporting_period_columns_unaffected_by_unpivot_feature(self):
        # reporting_period_columns drives dbt presentation provisioning's
        # revision-selection logic and must stay independent of the new
        # unpivot reshape feature.
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            reporting_period_columns=["Period"],
        )
        ((name, content, h, metrics),) = normalize_payload_to_csv(
            "data.csv", CSV_SIMPLE, target
        )
        assert content == CSV_SIMPLE
        assert target.reporting_period_columns == ["Period"]
        assert target.unpivot is None

    def test_zip_member_content_hash_is_of_extracted_content(self):
        import hashlib

        zip_payload = _make_zip(("x.csv", CSV_SIMPLE))
        ((_, content, h, _),) = normalize_payload_to_csv("archive.zip", zip_payload)
        assert h == hashlib.sha256(content).hexdigest()

    def test_source_and_normalized_bytes_in_metrics(self):
        zip_payload = _make_zip(("x.csv", CSV_SIMPLE))
        ((_, content, _, metrics),) = normalize_payload_to_csv(
            "archive.zip", zip_payload
        )
        assert metrics["source_bytes"] == len(zip_payload)
        assert metrics["normalized_bytes"] == len(content)

    def test_csv_filename_preserved_with_directory_stripped(self):
        ((name, _, _, _),) = normalize_payload_to_csv("subdir/data.csv", CSV_SIMPLE)
        # Should preserve original source_name
        assert "data.csv" in name

    def test_single_output_excel_content_hash_matches_source_payload(self):
        import hashlib

        xlsx = _make_xlsx(["col"], [["val1"]])
        ((_, _, h, _),) = normalize_payload_to_csv("report.xlsx", xlsx)
        assert h == hashlib.sha256(xlsx).hexdigest()

    def test_sheet_splitting_produces_distinct_hashes_per_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Jan"
        wb.active.append(["a"])
        wb.active.append([1])
        feb = wb.create_sheet("Feb")
        feb.append(["a"])
        feb.append([2])
        buf = io.BytesIO()
        wb.save(buf)

        sub_table = SubTableConfig(
            object_name_suffix="MONTHLY",
            adls_path_prefix="dataset/monthly",
            sheet_name_patterns=[".*"],
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            sub_tables=[sub_table],
        )
        results = normalize_payload_to_csv("report.xlsx", buf.getvalue(), target)
        assert len(results) == 2
        hashes = {h for _, _, h, _ in results}
        assert len(hashes) == 2
        for _, _, _, metrics in results:
            assert metrics["matched_sub_table_adls_path_prefix"] == "dataset/monthly"

    def test_zip_containing_excel_with_sheet_splitting(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        wb.active.append(["junk"])
        data_sheet = wb.create_sheet("Data")
        data_sheet.append(["a", "b"])
        data_sheet.append([1, 2])
        buf = io.BytesIO()
        wb.save(buf)
        zip_payload = _make_zip(("report.xlsx", buf.getvalue()))

        sub_table = SubTableConfig(
            object_name_suffix="DATA",
            adls_path_prefix="dataset/data",
            sheet_name_patterns=["^Data"],
        )
        target = TargetConfig(
            sub_dataset_id="t",
            object_name_suffix="T",
            adls_path_prefix="t",
            sub_tables=[sub_table],
        )
        ((name, content, _, metrics),) = normalize_payload_to_csv(
            "archive.zip", zip_payload, target
        )
        assert name == "report__Data.csv"
        assert metrics["source_file_type"] == "zip"
        assert metrics["extracted_from_archive"] is True
        assert metrics["matched_sub_table_adls_path_prefix"] == "dataset/data"


# ---------------------------------------------------------------------------
# resolve_sub_table_adls_prefix
# ---------------------------------------------------------------------------


def _make_target(sub_tables=None) -> TargetConfig:
    return TargetConfig(
        sub_dataset_id="test-target",
        object_name_suffix="TEST_TARGET",
        adls_path_prefix="dataset/test-target",
        sub_tables=sub_tables or [],
    )


def _make_sub_table(suffix, prefix, patterns) -> SubTableConfig:
    return SubTableConfig(
        object_name_suffix=suffix,
        adls_path_prefix=prefix,
        filename_patterns=patterns,
    )


class TestResolveSubTableAdlsPrefix:
    def test_no_sub_tables_returns_parent_prefix(self):
        target = _make_target()
        assert (
            resolve_sub_table_adls_prefix("DATA_FILE_Jan26.csv", {}, target)
            == "dataset/test-target"
        )

    def test_matching_filename_returns_sub_table_prefix(self):
        st = _make_sub_table(
            "TEST_TARGET_COVERAGE", "dataset/test-target-coverage", ["COVERAGE_FILE"]
        )
        target = _make_target([st])
        assert (
            resolve_sub_table_adls_prefix("COVERAGE_FILE_Jan26.csv", {}, target)
            == "dataset/test-target-coverage"
        )

    def test_non_matching_filename_returns_parent_prefix(self):
        st = _make_sub_table(
            "TEST_TARGET_COVERAGE", "dataset/test-target-coverage", ["COVERAGE_FILE"]
        )
        target = _make_target([st])
        assert (
            resolve_sub_table_adls_prefix("DATA_FILE_Jan26.csv", {}, target)
            == "dataset/test-target"
        )

    def test_pattern_match_is_case_insensitive(self):
        st = _make_sub_table(
            "TEST_TARGET_COVERAGE", "dataset/test-target-coverage", ["coverage_file"]
        )
        target = _make_target([st])
        assert (
            resolve_sub_table_adls_prefix("COVERAGE_FILE.csv", {}, target)
            == "dataset/test-target-coverage"
        )

    def test_first_matching_sub_table_wins(self):
        st1 = _make_sub_table("TEST_TARGET_A", "dataset/test-target-a", ["FILE_A"])
        st2 = _make_sub_table("TEST_TARGET_B", "dataset/test-target-b", ["FILE_A"])
        target = _make_target([st1, st2])
        assert (
            resolve_sub_table_adls_prefix("FILE_A.csv", {}, target)
            == "dataset/test-target-a"
        )

    def test_second_pattern_in_array_matches(self):
        st = _make_sub_table(
            "TEST_TARGET_COVERAGE",
            "dataset/test-target-coverage",
            ["COVERAGE_PATTERN_V1", "COVERAGE_PATTERN_V2"],
        )
        target = _make_target([st])
        assert (
            resolve_sub_table_adls_prefix("COVERAGE_PATTERN_V2_Jan26.csv", {}, target)
            == "dataset/test-target-coverage"
        )

    def test_anchor_pattern_only_matches_start(self):
        st = _make_sub_table(
            "TEST_TARGET_MAPPING", "dataset/test-target-mapping", ["^Mapping"]
        )
        target = _make_target([st])
        assert (
            resolve_sub_table_adls_prefix("Mapping.csv", {}, target)
            == "dataset/test-target-mapping"
        )
        assert (
            resolve_sub_table_adls_prefix("NotMapping.csv", {}, target)
            == "dataset/test-target"
        )

    def test_sheet_routed_metrics_short_circuit_filename_matching(self):
        st = _make_sub_table(
            "TEST_TARGET_COVERAGE", "dataset/test-target-coverage", ["COVERAGE_FILE"]
        )
        target = _make_target([st])
        metrics = {"matched_sub_table_adls_path_prefix": "dataset/sheet-routed"}
        assert (
            resolve_sub_table_adls_prefix(
                "DOES_NOT_MATCH_ANY_PATTERN.csv", metrics, target
            )
            == "dataset/sheet-routed"
        )
